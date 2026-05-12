"""Department of Education CCS public market benchmark parsing.

This module intentionally treats CCS quarterly data as public aggregate market
evidence. It does not infer target occupancy, waitlists, revenue, EBITDA, or
licensed-place capacity from CCS workbooks.
"""

from __future__ import annotations

from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
import re
from typing import Any

import openpyxl


SOURCE_NAME = "Department of Education CCS quarterly data"
SOURCE_QUALITY = "authoritative_public_aggregate"

CCS_CAVEATS = [
    "CCS data is public aggregate market evidence, not target-level evidence.",
    "Children using care measures realised CCS usage, not total latent demand.",
    "SA3 metrics are macro-local benchmarks, not micro-catchment proof.",
    "Approved services are operating services, not licensed places or actual vacancies.",
    "CBDC service count is a service-count proxy, not approved-place capacity.",
    "High fees may indicate pricing power or affordability / gap-fee pressure.",
    "Metrics are as of the published CCS quarter and may be stale.",
]

UNDERWRITING_USE = [
    "market_depth_benchmark",
    "cbdc_pricing_benchmark",
    "fee_cap_pressure_screen",
    "new_entrant_plausibility_context",
]

NOT_UNDERWRITING_USE = [
    "target_occupancy",
    "target_waitlist",
    "target_revenue",
    "target_ebitda",
    "definitive_unmet_demand",
    "licensed_place_capacity",
    "actual_vacancies",
]


def _checksum(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("%", " pct ")
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _parse_number(value: Any) -> float | int | None:
    if value is None or value == ".":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    text = str(value).strip()
    if not text or text == ".":
        return None
    text = text.replace("$", "").replace(",", "").replace("%", "").strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _safe_div(numerator: Any, denominator: Any) -> float | None:
    n = _parse_number(numerator)
    d = _parse_number(denominator)
    if n is None or d in (None, 0):
        return None
    return float(n) / float(d)


def _find_header_row(sheet: Any, required_tokens: list[str]) -> tuple[int, dict[str, int]]:
    required = [_normalise_header(token) for token in required_tokens]
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        headers = {_normalise_header(value): index for index, value in enumerate(row) if value is not None}
        header_text = " ".join(headers)
        if all(token in header_text for token in required):
            return row_index, headers
    raise ValueError(f"Could not find header row in sheet {sheet.title!r}")


def _column(headers: dict[str, int], *patterns: str) -> int | None:
    for pattern in patterns:
        regex = re.compile(pattern)
        for header, index in headers.items():
            if regex.search(header):
                return index
    return None


def _value(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _metric_base(quarter: str, sa3_code: str | None, sa3_name: str) -> dict[str, Any]:
    return {
        "quarter": quarter,
        "sa3_code": sa3_code,
        "sa3_name": sa3_name,
        "state": None,
        "children_0_5_using_care": None,
        "children_6_plus_using_care": None,
        "total_children_using_care": None,
        "families_using_care": None,
        "all_approved_services": None,
        "cbdc_services": None,
        "children_0_5_per_cbdc_service": None,
        "total_children_per_all_service": None,
        "cbdc_density_per_1000_children_0_5": None,
        "cbdc_mean_fee_per_hour": None,
        "cbdc_fee_growth_yoy_pct": None,
        "cbdc_services_above_cap_pct": None,
        "caveats": list(CCS_CAVEATS),
        "underwriting_use": list(UNDERWRITING_USE),
        "not_underwriting_use": list(NOT_UNDERWRITING_USE),
    }


def _parse_statistical_area(sheet: Any, quarter: str) -> dict[str, dict[str, Any]]:
    header_row, headers = _find_header_row(sheet, ["sa3 code", "sa3 name"])
    code_col = _column(headers, r"\bsa3 code\b")
    name_col = _column(headers, r"\bsa3 name\b")
    children_0_5_col = _column(headers, r"children.*0.*5")
    children_6_plus_col = _column(headers, r"children.*6")
    total_children_col = _column(headers, r"children.*total")
    families_col = _column(headers, r"\bfamilies\b")
    services_col = _column(headers, r"\bservices\b")
    if code_col is None or name_col is None:
        raise ValueError("Statistical Area sheet is missing SA3 code/name columns")

    metrics: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        sa3_name = str(_value(row, name_col) or "").strip()
        if not sa3_name:
            continue
        raw_code = _value(row, code_col)
        sa3_code = str(raw_code).strip() if raw_code is not None else None
        metric = _metric_base(quarter, sa3_code, sa3_name)
        metric["children_0_5_using_care"] = _parse_number(_value(row, children_0_5_col))
        metric["children_6_plus_using_care"] = _parse_number(_value(row, children_6_plus_col))
        metric["total_children_using_care"] = _parse_number(_value(row, total_children_col))
        metric["families_using_care"] = _parse_number(_value(row, families_col))
        metric["all_approved_services"] = _parse_number(_value(row, services_col))
        metric["total_children_per_all_service"] = _safe_div(metric["total_children_using_care"], metric["all_approved_services"])
        metrics[sa3_code or sa3_name.lower()] = metric
    return metrics


def _parse_cbdc_fees(sheet: Any) -> dict[str, dict[str, Any]]:
    header_row, headers = _find_header_row(sheet, ["sa3 code", "sa3 name"])
    code_col = _column(headers, r"\bsa3 code\b")
    name_col = _column(headers, r"\bsa3 name\b")
    state_col = _column(headers, r"\bstate\b")
    services_col = _column(headers, r"service count")
    fee_col = _column(headers, r"mean fee.*hour")
    growth_col = _column(headers, r"growth.*mean fee")
    above_cap_col = _column(headers, r"pct services.*above.*cap")
    if above_cap_col is None:
        above_cap_col = _column(headers, r"services.*above.*cap$")
    if code_col is None or name_col is None:
        raise ValueError("CBDC Fees sheet is missing SA3 code/name columns")

    rows: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        raw_code = _value(row, code_col)
        sa3_name = str(_value(row, name_col) or "").strip()
        if raw_code is None and not sa3_name:
            continue
        sa3_code = str(raw_code).strip() if raw_code is not None else None
        rows[sa3_code or sa3_name.lower()] = {
            "state": str(_value(row, state_col) or "").strip() or None,
            "cbdc_services": _parse_number(_value(row, services_col)),
            "cbdc_mean_fee_per_hour": _parse_number(_value(row, fee_col)),
            "cbdc_fee_growth_yoy_pct": _parse_number(_value(row, growth_col)),
            "cbdc_services_above_cap_pct": _parse_number(_value(row, above_cap_col)),
        }
    return rows


def _round(value: float | None) -> float | None:
    return None if value is None else round(value, 4)


def parse_ccs_workbook(path: str, quarter: str, source_url: str | None = None) -> dict[str, Any]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(path)

    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - exception type depends on openpyxl internals
        raise ValueError(f"Could not open CCS workbook: {exc}") from exc

    required_tabs = ["Statistical Area", "CBDC Fees"]
    missing_tabs = [name for name in required_tabs if name not in workbook.sheetnames]
    if missing_tabs:
        raise ValueError(f"CCS workbook missing required tab(s): {', '.join(missing_tabs)}")

    warnings: list[str] = []
    metrics = _parse_statistical_area(workbook["Statistical Area"], quarter)
    cbdc_rows = _parse_cbdc_fees(workbook["CBDC Fees"])

    for key, metric in metrics.items():
        cbdc = cbdc_rows.get(key)
        if not cbdc:
            warnings.append(f"CBDC Fees row not found for SA3 {metric['sa3_code'] or metric['sa3_name']}; CBDC metrics left null.")
            continue
        metric.update(cbdc)
        if not metric.get("state"):
            warnings.append(f"State could not be joined from CBDC Fees for SA3 {metric['sa3_code'] or metric['sa3_name']}; state left null.")
        metric["children_0_5_per_cbdc_service"] = _round(_safe_div(metric["children_0_5_using_care"], metric["cbdc_services"]))
        density = _safe_div(metric["cbdc_services"], metric["children_0_5_using_care"])
        metric["cbdc_density_per_1000_children_0_5"] = _round(density * 1000 if density is not None else None)
        metric["total_children_per_all_service"] = _round(metric.get("total_children_per_all_service"))

    return {
        "version": {
            "quarter": quarter,
            "source_url": source_url,
            "filename": workbook_path.name,
            "checksum_sha256": _checksum(workbook_path),
            "ingested_at": datetime.now(timezone.utc).isoformat(),
            "source": SOURCE_NAME,
            "source_quality": SOURCE_QUALITY,
        },
        "sa3_metrics": list(metrics.values()),
        "warnings": warnings,
    }


def get_sa3_ccs_metric(parsed: dict[str, Any], sa3_name: str | None = None, sa3_code: str | None = None) -> dict[str, Any] | None:
    code = str(sa3_code).strip() if sa3_code is not None else None
    name = sa3_name.strip().lower() if sa3_name else None
    for metric in parsed.get("sa3_metrics", []):
        if code and str(metric.get("sa3_code")) == code:
            return metric
        if name and str(metric.get("sa3_name", "")).strip().lower() == name:
            return metric
    return None


def build_public_market_benchmark(
    parsed_ccs: dict[str, Any] | None,
    target_sa3_code: str | None = None,
    target_sa3_name: str | None = None,
    sa3_selection_source: str | None = None,
) -> dict[str, Any] | None:
    """Format a matched SA3 CCS row for workflow.market_audit attachment."""
    if not isinstance(parsed_ccs, dict):
        return None
    if not target_sa3_code and not target_sa3_name:
        return None
    metric = get_sa3_ccs_metric(parsed_ccs, sa3_name=target_sa3_name, sa3_code=target_sa3_code)
    if not metric:
        return None
    version = parsed_ccs.get("version") if isinstance(parsed_ccs.get("version"), dict) else {}
    caveats = list(metric.get("caveats") or CCS_CAVEATS)
    if sa3_selection_source == "manual_context":
        caveats.append("SA3 selection came from manual/admin context, not source-document extraction.")
    benchmark = {
        "source": version.get("source") or SOURCE_NAME,
        "source_quality": version.get("source_quality") or SOURCE_QUALITY,
        "as_of_quarter": version.get("quarter") or metric.get("quarter"),
        "sa3_code": metric.get("sa3_code"),
        "sa3_name": metric.get("sa3_name"),
        "state": metric.get("state"),
        "children_0_5_using_care": metric.get("children_0_5_using_care"),
        "children_6_plus_using_care": metric.get("children_6_plus_using_care"),
        "total_children_using_care": metric.get("total_children_using_care"),
        "families_using_care": metric.get("families_using_care"),
        "all_approved_services": metric.get("all_approved_services"),
        "cbdc_services": metric.get("cbdc_services"),
        "children_0_5_per_cbdc_service": metric.get("children_0_5_per_cbdc_service"),
        "total_children_per_all_service": metric.get("total_children_per_all_service"),
        "cbdc_density_per_1000_children_0_5": metric.get("cbdc_density_per_1000_children_0_5"),
        "cbdc_mean_fee_per_hour": metric.get("cbdc_mean_fee_per_hour"),
        "cbdc_fee_growth_yoy_pct": metric.get("cbdc_fee_growth_yoy_pct"),
        "cbdc_services_above_cap_pct": metric.get("cbdc_services_above_cap_pct"),
        "caveats": caveats,
        "underwriting_use": list(metric.get("underwriting_use") or UNDERWRITING_USE),
        "not_underwriting_use": list(metric.get("not_underwriting_use") or NOT_UNDERWRITING_USE),
    }
    if sa3_selection_source:
        benchmark["sa3_selection_source"] = sa3_selection_source
        benchmark["sa3_selection_note"] = (
            "SA3 was supplied through manual/admin context."
            if sa3_selection_source == "manual_context"
            else "SA3 was explicitly extracted from source documents."
        )
    return benchmark


def attach_ccs_public_market_benchmark_if_available(
    market_audit: dict[str, Any],
    parsed_ccs: dict[str, Any] | None,
    target_sa3_code: str | None = None,
    target_sa3_name: str | None = None,
    sa3_selection_source: str | None = None,
) -> dict[str, Any]:
    """Attach a CCS public aggregate benchmark when an explicit target SA3 is supplied."""
    result = dict(market_audit or {})
    benchmark = build_public_market_benchmark(
        parsed_ccs,
        target_sa3_code=target_sa3_code,
        target_sa3_name=target_sa3_name,
        sa3_selection_source=sa3_selection_source,
    )
    if benchmark:
        result["public_market_benchmark"] = benchmark
    return result


def rank_sa3_by_children_0_5_per_cbdc_service(parsed: dict[str, Any], state: str | None = None) -> list[dict[str, Any]]:
    state_filter = state.strip().lower() if state else None
    metrics = []
    for metric in parsed.get("sa3_metrics", []):
        if state_filter and str(metric.get("state", "")).lower() != state_filter:
            continue
        if metric.get("children_0_5_per_cbdc_service") is not None:
            metrics.append(metric)
    return sorted(metrics, key=lambda item: item["children_0_5_per_cbdc_service"], reverse=True)
