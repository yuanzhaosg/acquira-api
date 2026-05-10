import os
import tempfile
import unittest
from pathlib import Path

import openpyxl

from ccs_market_data import (
    CCS_CAVEATS,
    attach_ccs_public_market_benchmark_if_available,
    build_public_market_benchmark,
    get_sa3_ccs_metric,
    parse_ccs_workbook,
    rank_sa3_by_children_0_5_per_cbdc_service,
)


FORBIDDEN_PHRASES = [
    "demand per centre",
    "proof of demand",
    "proof of occupancy",
    "actual demand",
    "true demand",
    "definitive unmet demand",
]


def make_synthetic_workbook(path: Path, include_state: bool = True) -> None:
    workbook = openpyxl.Workbook()
    statistical = workbook.active
    statistical.title = "Statistical Area"
    statistical.append(["Table 8.1: Number of children, families and services by SA3, December quarter 2025"])
    statistical.append([])
    statistical.append([
        "SA3 code",
        "SA3 name",
        "Number of children -\n0 to 5 years",
        "Number of children -\n6+ years",
        "Number of children -\nTotal",
        "Number of families",
        "Number of services",
    ])
    statistical.append([21104, "Whitehorse - East", 2320, 1390, 3710, 2840, 38])
    statistical.append([29999, "Synthetic Growth", 5000, 1000, 6000, 3900, 40])

    cbdc = workbook.create_sheet("CBDC Fees")
    cbdc.append(["Table 9.1: Centre Based Day Care fee per hour analysis by SA3"])
    cbdc.append([])
    cbdc.append([
        "SA4 Code",
        "SA4 Name",
        "State",
        "SA3 code",
        "SA3 Name",
        "Dec 2025\nService count",
        "Dec 2025\nMean fee\nper hour",
        "% Growth in\nMean fee (Dec 2024 to Dec 2025)",
        "Dec 2025\nNumber of services above cap",
        "% Services\nabove the cap",
    ])
    cbdc.append([211, "Melbourne - Outer East", "VIC" if include_state else None, 21104, "Whitehorse - East", 28, 14.52, 4.68812983504037, 13, 46.4285714285714])
    cbdc.append([299, "Synthetic", "VIC", 29999, "Synthetic Growth", 20, 15.50, 6.25, 8, 40])

    # Present to ensure OSHC data is not accidentally used as the CBDC benchmark.
    oshc = workbook.create_sheet("OSHC Fees")
    oshc.append(["Table 10.1: Outside School Hours Care fee per hour analysis by SA3"])
    oshc.append([])
    oshc.append(["SA3 code", "SA3 Name", "Dec 2025\nMean fee\nper hour"])
    oshc.append([21104, "Whitehorse - East", 99.99])
    workbook.save(path)


class CcsMarketDataTests(unittest.TestCase):
    def parse_synthetic(self, include_state: bool = True) -> dict:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "ccs-synthetic.xlsx"
            make_synthetic_workbook(path, include_state=include_state)
            return parse_ccs_workbook(str(path), "Dec 2025", source_url="https://example.test/ccs.xlsx")

    def test_parse_whitehorse_east_synthetic_values(self):
        parsed = self.parse_synthetic()
        metric = get_sa3_ccs_metric(parsed, sa3_code="21104")
        self.assertIsNotNone(metric)
        assert metric is not None
        self.assertEqual(metric["sa3_name"], "Whitehorse - East")
        self.assertEqual(metric["state"], "VIC")
        self.assertEqual(metric["children_0_5_using_care"], 2320)
        self.assertEqual(metric["children_6_plus_using_care"], 1390)
        self.assertEqual(metric["total_children_using_care"], 3710)
        self.assertEqual(metric["families_using_care"], 2840)
        self.assertEqual(metric["all_approved_services"], 38)
        self.assertEqual(metric["cbdc_services"], 28)
        self.assertAlmostEqual(metric["cbdc_mean_fee_per_hour"], 14.52, places=2)
        self.assertAlmostEqual(metric["cbdc_fee_growth_yoy_pct"], 4.69, places=2)
        self.assertAlmostEqual(metric["cbdc_services_above_cap_pct"], 46.4, places=1)
        self.assertAlmostEqual(metric["children_0_5_per_cbdc_service"], 82.8571, places=3)
        self.assertAlmostEqual(metric["total_children_per_all_service"], 97.6316, places=3)
        self.assertAlmostEqual(metric["cbdc_density_per_1000_children_0_5"], 12.069, places=3)
        self.assertEqual(parsed["version"]["source"], "Department of Education CCS quarterly data")
        self.assertEqual(parsed["version"]["source_quality"], "authoritative_public_aggregate")
        self.assertIn("warnings", parsed)

    def test_primary_cbdc_metric_is_not_total_children_per_all_services(self):
        parsed = self.parse_synthetic()
        metric = get_sa3_ccs_metric(parsed, sa3_name="Whitehorse - East")
        self.assertIsNotNone(metric)
        assert metric is not None
        self.assertNotEqual(metric["children_0_5_per_cbdc_service"], metric["total_children_per_all_service"])
        self.assertAlmostEqual(metric["children_0_5_per_cbdc_service"], 2320 / 28, places=4)

    def test_oshc_fee_data_is_not_used_as_cbdc_pricing(self):
        parsed = self.parse_synthetic()
        metric = get_sa3_ccs_metric(parsed, sa3_code="21104")
        self.assertIsNotNone(metric)
        assert metric is not None
        self.assertEqual(metric["cbdc_mean_fee_per_hour"], 14.52)
        self.assertNotEqual(metric["cbdc_mean_fee_per_hour"], 99.99)

    def test_semantics_do_not_label_ccs_as_target_evidence_or_capacity(self):
        parsed = self.parse_synthetic()
        metric = get_sa3_ccs_metric(parsed, sa3_code="21104")
        self.assertIsNotNone(metric)
        assert metric is not None
        self.assertIn("target_occupancy", metric["not_underwriting_use"])
        self.assertIn("target_waitlist", metric["not_underwriting_use"])
        self.assertIn("target_revenue", metric["not_underwriting_use"])
        self.assertIn("target_ebitda", metric["not_underwriting_use"])
        self.assertIn("licensed_place_capacity", metric["not_underwriting_use"])
        self.assertIn("actual_vacancies", metric["not_underwriting_use"])
        self.assertIn("market_depth_benchmark", metric["underwriting_use"])
        self.assertIn("cbdc_pricing_benchmark", metric["underwriting_use"])
        caveat_text = " ".join(metric["caveats"])
        self.assertIn("public aggregate market evidence", caveat_text)
        self.assertIn("not target-level evidence", caveat_text)
        self.assertIn("not licensed places or actual vacancies", caveat_text)

    def test_forbidden_phrases_do_not_appear_in_caveats_or_labels(self):
        parsed = self.parse_synthetic()
        metric = get_sa3_ccs_metric(parsed, sa3_code="21104")
        assert metric is not None
        label_text = " ".join(str(key) for key in metric.keys()).lower()
        caveat_text = " ".join(metric["caveats"]).lower()
        for phrase in FORBIDDEN_PHRASES:
            self.assertNotIn(phrase, label_text)
            self.assertNotIn(phrase, caveat_text)
        self.assertFalse(any(phrase in " ".join(CCS_CAVEATS).lower() for phrase in FORBIDDEN_PHRASES))

    def test_missing_state_join_warns_and_leaves_state_null(self):
        parsed = self.parse_synthetic(include_state=False)
        metric = get_sa3_ccs_metric(parsed, sa3_code="21104")
        self.assertIsNotNone(metric)
        assert metric is not None
        self.assertIsNone(metric["state"])
        self.assertTrue(any("State could not be joined" in warning for warning in parsed["warnings"]))

    def test_rank_by_children_0_5_per_cbdc_service(self):
        parsed = self.parse_synthetic()
        ranked = rank_sa3_by_children_0_5_per_cbdc_service(parsed, state="VIC")
        self.assertGreaterEqual(len(ranked), 2)
        self.assertEqual(ranked[0]["sa3_name"], "Synthetic Growth")
        self.assertGreater(ranked[0]["children_0_5_per_cbdc_service"], ranked[1]["children_0_5_per_cbdc_service"])

    def test_public_market_benchmark_requires_parsed_ccs_and_target_sa3(self):
        self.assertIsNone(build_public_market_benchmark(None, target_sa3_code="21104"))
        parsed = self.parse_synthetic()
        self.assertIsNone(build_public_market_benchmark(parsed))
        self.assertIsNone(build_public_market_benchmark(parsed, target_sa3_code="00000"))

    def test_public_market_benchmark_builds_from_matching_sa3_code(self):
        parsed = self.parse_synthetic()
        benchmark = build_public_market_benchmark(parsed, target_sa3_code="21104")
        self.assertIsNotNone(benchmark)
        assert benchmark is not None
        self.assertEqual(benchmark["source"], "Department of Education CCS quarterly data")
        self.assertEqual(benchmark["source_quality"], "authoritative_public_aggregate")
        self.assertEqual(benchmark["as_of_quarter"], "Dec 2025")
        self.assertEqual(benchmark["sa3_code"], "21104")
        self.assertEqual(benchmark["sa3_name"], "Whitehorse - East")
        self.assertEqual(benchmark["state"], "VIC")
        self.assertEqual(benchmark["children_0_5_using_care"], 2320)
        self.assertEqual(benchmark["children_6_plus_using_care"], 1390)
        self.assertEqual(benchmark["total_children_using_care"], 3710)
        self.assertEqual(benchmark["families_using_care"], 2840)
        self.assertEqual(benchmark["all_approved_services"], 38)
        self.assertEqual(benchmark["cbdc_services"], 28)
        self.assertAlmostEqual(benchmark["children_0_5_per_cbdc_service"], 82.86, places=2)
        self.assertAlmostEqual(benchmark["total_children_per_all_service"], 97.63, places=2)
        self.assertAlmostEqual(benchmark["cbdc_density_per_1000_children_0_5"], 12.07, places=2)
        self.assertAlmostEqual(benchmark["cbdc_mean_fee_per_hour"], 14.52, places=2)
        self.assertAlmostEqual(benchmark["cbdc_fee_growth_yoy_pct"], 4.69, places=2)
        self.assertAlmostEqual(benchmark["cbdc_services_above_cap_pct"], 46.4, places=1)
        self.assertIn("public aggregate market evidence", " ".join(benchmark["caveats"]))
        self.assertIn("target_occupancy", benchmark["not_underwriting_use"])
        self.assertIn("licensed_place_capacity", benchmark["not_underwriting_use"])

    def test_public_market_benchmark_builds_from_matching_sa3_name(self):
        parsed = self.parse_synthetic()
        benchmark = build_public_market_benchmark(parsed, target_sa3_name="whitehorse - east")
        self.assertIsNotNone(benchmark)
        assert benchmark is not None
        self.assertEqual(benchmark["sa3_code"], "21104")
        self.assertAlmostEqual(benchmark["children_0_5_per_cbdc_service"], 82.86, places=2)

    def test_attach_ccs_public_market_benchmark_is_noop_until_matched(self):
        parsed = self.parse_synthetic()
        audit = {"warnings": ["existing warning"]}
        self.assertEqual(attach_ccs_public_market_benchmark_if_available(audit, None, target_sa3_code="21104"), audit)
        self.assertEqual(attach_ccs_public_market_benchmark_if_available(audit, parsed), audit)
        self.assertEqual(attach_ccs_public_market_benchmark_if_available(audit, parsed, target_sa3_code="00000"), audit)

    def test_attach_ccs_public_market_benchmark_adds_only_benchmark(self):
        parsed = self.parse_synthetic()
        audit = {"warnings": ["existing warning"]}
        attached = attach_ccs_public_market_benchmark_if_available(audit, parsed, target_sa3_code="21104")
        self.assertEqual(attached["warnings"], ["existing warning"])
        self.assertIn("public_market_benchmark", attached)
        self.assertNotIn("local_demand_supply", attached)
        self.assertNotIn("public_market_benchmark", audit)

    def test_public_market_benchmark_avoids_forbidden_phrases(self):
        parsed = self.parse_synthetic()
        benchmark = build_public_market_benchmark(parsed, target_sa3_code="21104")
        assert benchmark is not None
        text = " ".join(str(value).lower() for value in benchmark.values() if not isinstance(value, list))
        text += " " + " ".join(str(item).lower() for key in ("caveats", "underwriting_use") for item in benchmark[key])
        for phrase in FORBIDDEN_PHRASES:
            self.assertNotIn(phrase, text)

    def test_missing_required_tabs_hard_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "bad.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "Summary"
            workbook.save(path)
            with self.assertRaises(ValueError):
                parse_ccs_workbook(str(path), "Dec 2025")

    @unittest.skipUnless(os.environ.get("ACQUIRA_REAL_CCS_WORKBOOK"), "Set ACQUIRA_REAL_CCS_WORKBOOK to run optional real workbook QA.")
    def test_optional_real_dec_2025_whitehorse_east_workbook(self):
        parsed = parse_ccs_workbook(os.environ["ACQUIRA_REAL_CCS_WORKBOOK"], "Dec 2025")
        metric = get_sa3_ccs_metric(parsed, sa3_code="21104")
        self.assertIsNotNone(metric)
        assert metric is not None
        self.assertEqual(metric["children_0_5_using_care"], 2320)
        self.assertEqual(metric["children_6_plus_using_care"], 1390)
        self.assertEqual(metric["total_children_using_care"], 3710)
        self.assertEqual(metric["families_using_care"], 2840)
        self.assertEqual(metric["all_approved_services"], 38)
        self.assertEqual(metric["cbdc_services"], 28)
        self.assertAlmostEqual(metric["cbdc_mean_fee_per_hour"], 14.52, places=2)
        self.assertAlmostEqual(metric["cbdc_fee_growth_yoy_pct"], 4.69, places=2)
        self.assertAlmostEqual(metric["cbdc_services_above_cap_pct"], 46.4, places=1)
        self.assertAlmostEqual(metric["children_0_5_per_cbdc_service"], 82.86, places=2)


if __name__ == "__main__":
    unittest.main()
