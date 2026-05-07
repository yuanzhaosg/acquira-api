import asyncio
import importlib.util
import os
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("ANTHROPIC_API_KEY", "test-key")
os.environ.setdefault("SUPABASE_URL", "https://example.supabase.co")
os.environ.setdefault("SUPABASE_SERVICE_KEY", "test-service-key")

if importlib.util.find_spec("fastapi") is None:
    fastapi = types.ModuleType("fastapi")

    class FastAPI:
        def __init__(self, *args, **kwargs):
            pass

        def add_middleware(self, *args, **kwargs):
            pass

        def get(self, *args, **kwargs):
            return lambda fn: fn

        def post(self, *args, **kwargs):
            return lambda fn: fn

    fastapi.FastAPI = FastAPI
    fastapi.Query = lambda *args, **kwargs: None
    middleware = types.ModuleType("fastapi.middleware")
    cors = types.ModuleType("fastapi.middleware.cors")
    cors.CORSMiddleware = object
    responses = types.ModuleType("fastapi.responses")
    responses.StreamingResponse = object
    responses.JSONResponse = lambda *args, **kwargs: {"args": args, "kwargs": kwargs}
    sys.modules["fastapi"] = fastapi
    sys.modules["fastapi.middleware"] = middleware
    sys.modules["fastapi.middleware.cors"] = cors
    sys.modules["fastapi.responses"] = responses

if importlib.util.find_spec("anthropic") is None:
    anthropic = types.ModuleType("anthropic")
    anthropic.Anthropic = lambda *args, **kwargs: types.SimpleNamespace(messages=types.SimpleNamespace(create=None))
    sys.modules["anthropic"] = anthropic

if importlib.util.find_spec("supabase") is None:
    supabase = types.ModuleType("supabase")
    supabase.create_client = lambda *args, **kwargs: types.SimpleNamespace()
    sys.modules["supabase"] = supabase

if importlib.util.find_spec("xlrd") is None:
    xlrd = types.ModuleType("xlrd")
    xlrd.open_workbook = lambda *args, **kwargs: None
    sys.modules["xlrd"] = xlrd

import fitz
import openpyxl

from structured_deal import build_structured_deal_intelligence

import main


def make_mixed_pdf(path: Path) -> None:
    image_doc = fitz.open()
    image_page = image_doc.new_page(width=420, height=220)
    image_page.insert_text((24, 34), "Revenue 1,200,000", fontsize=14)
    image_page.insert_text((24, 64), "Wages 620,000", fontsize=14)
    image_page.insert_text((24, 94), "Rent 110,000", fontsize=14)
    image_page.insert_text((24, 124), "EBITDA 210,000", fontsize=14)
    pix = image_page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)

    doc = fitz.open()
    normal = doc.new_page(width=595, height=842)
    normal.insert_text((72, 72), "Information Memorandum\nCentre overview and lease summary.", fontsize=12)
    pl_page = doc.new_page(width=595, height=842)
    pl_page.insert_text((72, 72), "PROFIT AND LOSS", fontsize=18)
    pl_page.insert_image(fitz.Rect(72, 110, 520, 380), stream=pix.tobytes("png"))
    doc.save(path)
    doc.close()
    image_doc.close()


def make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adjusted Actuals"
    ws.append(["Metric", "FY25"])
    ws.append(["Revenue", 1200000])
    ws.append(["Rent", 110000])
    ws.append(["EBITDA", "=B2-B3-620000"])

    ws = wb.create_sheet("Output Occupancy")
    ws.append(["Month", "Occupancy"])
    ws.append(["Feb-26", 88.5])

    ws = wb.create_sheet("WorkedHours - Employment Hero")
    ws.append(["Role", "Hours", "Wages"])
    ws.append(["Educators", 1520, 620000])
    wb.save(path)


class DocumentExtractionTests(unittest.TestCase):
    def test_missing_anthropic_key_reports_missing_without_secret(self):
        original = os.environ.pop("ANTHROPIC_API_KEY", None)
        try:
            diagnostics = main.validate_vision_provider_config()
        finally:
            if original is not None:
                os.environ["ANTHROPIC_API_KEY"] = original

        self.assertFalse(diagnostics["configured"])
        self.assertEqual(diagnostics["auth_status"], "missing_api_key")
        self.assertNotIn("test-key", str(diagnostics))

    def test_vision_401_classifies_invalid_auth_without_key_value(self):
        error = "Error code: 401 - authentication_error invalid x-api-key for redacted-secret"
        self.assertEqual(main.classify_vision_provider_error(error), "invalid_auth")

    def test_pdf_page_level_fallback_is_invoked_for_sparse_pl_page(self):
        with tempfile.TemporaryDirectory() as tmp:
            pdf_path = Path(tmp) / "synthetic-im.pdf"
            make_mixed_pdf(pdf_path)

            async def fake_vision(_image_b64: str, filename: str, page_number: int, reason: str) -> str:
                return (
                    f"--- Page {page_number} ---\n"
                    f"SOURCE_FILE: {filename}\n"
                    f"PAGE: {page_number}\n"
                    "EXTRACTION_METHOD: pdf_vision\n"
                    f"FALLBACK_REASON: {reason}\n"
                    "| Revenue | 1,200,000 |\n"
                    "| Wages | 620,000 |\n"
                    "| Rent | 110,000 |\n"
                    "| EBITDA | 210,000 |"
                )

            with patch.object(main, "vision_extract_page", side_effect=fake_vision) as vision:
                text = asyncio.run(main.extract_pdf_pages_with_fallback(str(pdf_path), pdf_path.name, "im_pdf"))

            self.assertTrue(vision.called)
            self.assertIn("--- Page 2 ---", text)
            self.assertIn("EXTRACTION_METHOD: pdf_vision", text)
            self.assertIn("Revenue", text)
            self.assertIn("EBITDA", text)

    def test_workbook_digest_includes_sheets_coordinates_and_categories(self):
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = Path(tmp) / "synthetic-databook.xlsx"
            make_workbook(xlsx_path)

            digest = main.extract_excel_text(str(xlsx_path))

            self.assertIn("SHEET: Adjusted Actuals", digest)
            self.assertIn("DETECTED_CATEGORY: financials", digest)
            self.assertIn("SHEET: Output Occupancy", digest)
            self.assertIn("DETECTED_CATEGORY: occupancy", digest)
            self.assertIn("SHEET: WorkedHours - Employment Hero", digest)
            self.assertIn("DETECTED_CATEGORY: payroll_staffing", digest)
            self.assertIn("A1=Metric", digest)
            self.assertIn("B2=1200000", digest)
            self.assertIn("formula =B2-B3-620000", digest)
            self.assertIn("EVIDENCE_REFS: use synthetic-databook.xlsx / Adjusted Actuals / cell coordinates", digest)

    def test_market_audit_shell_exists_when_market_data_missing(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["synthetic-im.pdf"]},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {}},
                "occupancy": {},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text="Synthetic report without market data.",
            source_files=["synthetic-im.pdf"],
            file_classes={"synthetic-im.pdf": "im_pdf"},
        )

        self.assertIn("market_audit", workflow)
        self.assertEqual(workflow["market_audit"]["status"], "missing")
        self.assertIn("missing_fields", workflow["market_audit"])
        self.assertTrue(workflow["market_audit"]["warnings"])

    def test_provider_failure_warning_surfaces_in_workflow(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["synthetic-im.pdf"]},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {}},
                "occupancy": {},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "--- Page 12 ---\n"
                "SOURCE_FILE: synthetic-im.pdf\n"
                "PAGE: 12\n"
                "EXTRACTION_METHOD: pdf_vision\n"
                "VISION_PROVIDER_STATUS: invalid_auth\n"
                "VISION_EXTRACTION_ERROR: Error code: 401 - invalid x-api-key\n"
                "FALLBACK_REASON: sparse high-value evidence page 12\n"
            ),
            source_files=["synthetic-im.pdf"],
            file_classes={"synthetic-im.pdf": "im_pdf"},
        )

        messages = [warning["message"] for warning in workflow["extraction_warnings"]]
        self.assertTrue(any("pages 12" in message and "vision provider" in message for message in messages))
        self.assertNotIn("sk-ant", str(workflow["extraction_warnings"]))

    def test_workflow_tolerates_string_items_from_excel_extraction(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["databook.xlsx"], "missing_fields": "payroll detail"},
                "centre": "centre details not structured",
                "financials": {"fy25": "FY25 numbers were returned as text"},
                "occupancy": "occupancy returned as text",
                "key_ratios": "ratios returned as text",
                "hard_flags": ["labour costs require review"],
                "_market_audit": "market audit unavailable",
                "_pipeline_audit": "pipeline audit unavailable",
            },
            scored={
                "centre_name": "Synthetic Childcare",
                "deal_breaker_flags": {"flags": ["manual review required"]},
                "next_steps": {
                    "ask_broker_for": ["Source payroll summary"],
                    "due_diligence_priorities": ["Verify occupancy workbook tabs"],
                },
            },
            combined_text="=== databook.xlsx (pl_excel) ===\nWORKBOOK_DIGEST: databook.xlsx\nSHEET: Adjusted Actuals\nD23=Total revenue | E23=100",
            source_files=["databook.xlsx"],
            file_classes={"databook.xlsx": "pl_excel"},
        )

        self.assertIn("facts", workflow)
        self.assertTrue(any(risk["reason"] == "labour costs require review" for risk in workflow["risks"]))
        self.assertTrue(any(risk["reason"] == "manual review required" for risk in workflow["risks"]))
        self.assertIn("payroll detail", workflow["missing_fields"])

    def test_workbook_rows_derive_financial_and_occupancy_facts(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["databook.xlsx"], "missing_fields": ["revenue", "payroll", "ebitda", "occupancy_history"]},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {}},
                "occupancy": {},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== databook.xlsx (pl_excel) ===\n"
                "WORKBOOK_DIGEST: databook.xlsx\n"
                "SHEET: Adjusted Actuals\n"
                "D23=Total revenue | E23=100000 | F23=1200000\n"
                "D30=Total Employment Costs | E30=50000 | F30=620000\n"
                "D36=Rent | E36=9000 | F36=110000\n"
                "D41=Normalised EBITDA | E41=18000 | F41=210000\n"
                "SHEET: Output Occupancy\n"
                "A1=Week | B1=Occupancy\n"
                "A2=Week 1 | B2=0.82\n"
                "A3=Week 2 | B3=0.84\n"
                "A4=Week 3 | B4=0.86\n"
                "A5=Week 4 | B5=0.88\n"
            ),
            source_files=["databook.xlsx"],
            file_classes={"databook.xlsx": "pl_excel"},
        )

        fields = {fact["field"]: fact for fact in workflow["facts"]}
        self.assertEqual(fields["revenue"]["value"], 1200000)
        self.assertEqual(fields["payroll_labour_cost"]["value"], 620000)
        self.assertEqual(fields["rent_pa"]["value"], 110000)
        self.assertEqual(fields["normalised_ebitda"]["value"], 210000)
        self.assertEqual(fields["avg_4wk_occupancy_pct"]["value"], 85)
        self.assertEqual(fields["revenue"]["source_type"], "workbook_derived")
        self.assertEqual(fields["revenue"]["provenance"], "derived")
        self.assertEqual(fields["revenue"]["underwriting_use"], "review_required")
        self.assertIn(fields["revenue"]["trust"], {"medium", "disputed"})
        self.assertEqual(fields["revenue"]["source"]["cell_range"], "D23:F23")
        self.assertIn("Derived from workbook digest", fields["revenue"]["derivation_note"])
        self.assertIn("derivation_recipe", fields["avg_4wk_occupancy_pct"])
        self.assertIn("evidence_readiness", workflow)
        self.assertTrue(workflow["evidence_readiness"]["derived"])
        self.assertEqual(workflow["valuation_gate"]["required_evidence"]["occupancy_history"], True)
        self.assertEqual(workflow["valuation_gate"]["status"], "needs_review")
        self.assertNotIn("occupancy_history", [str(field).lower() for field in workflow["missing_fields"]])
        self.assertIn("canonical_facts", workflow)
        self.assertIn("valuation_gate_summary", workflow)
        self.assertIn("evidence_quality", workflow)
        self.assertIn("evidence_readiness", workflow)
        self.assertIn("partner_judgement_prompts", workflow)
        self.assertTrue(workflow["canonical_facts"].get("revenue"))
        self.assertTrue(workflow["valuation_gate_summary"].get("rows"))

    def test_workbook_derived_financials_prefer_workbook_and_flag_conflict(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["im.pdf", "databook.xlsx"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 900000, "ebitda": 100000, "total_labour_cost": 400000}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== databook.xlsx (pl_excel) ===\n"
                "SHEET: Adjusted Actuals\n"
                "D23=Total revenue | E23=1200000\n"
                "D30=Total Employment Costs | E30=620000\n"
                "D41=EBITDA | E41=210000\n"
            ),
            source_files=["im.pdf", "databook.xlsx"],
            file_classes={"im.pdf": "im_pdf", "databook.xlsx": "pl_excel"},
        )

        fields = {fact["field"]: fact for fact in workflow["facts"]}
        self.assertEqual(fields["revenue"]["value"], 1200000)
        self.assertEqual(fields["revenue"]["trust"], "disputed")
        self.assertEqual(fields["revenue"]["underwriting_use"], "review_required")
        self.assertTrue(fields["revenue"]["conflicts"])
        self.assertEqual(fields["payroll_labour_cost"]["value"], 620000)
        self.assertTrue(any(warning["id"] == "workbook_financial_conflicts" for warning in workflow["extraction_warnings"]))

    def test_supplemental_documents_populate_identity_fields(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["approval.pdf"], "missing_fields": ["trading_name", "address", "suburb", "postcode"]},
                "centre": {},
                "financials": {"fy25": {"revenue": 1000000, "ebitda": 200000, "total_labour_cost": 550000}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Fallback Centre"},
            combined_text=(
                "=== provider approval.pdf (service_approval_pdf) ===\n"
                "--- Page 2 ---\n"
                "Trading name: Jenny's Early Learning Norlane\n"
                "Service name: Jenny's Norlane Childcare\n"
                "Address: 12 Station Street, Norlane VIC 3214\n"
                "Suburb: Norlane\n"
                "Postcode: 3214\n"
            ),
            source_files=["approval.pdf"],
            file_classes={"approval.pdf": "service_approval_pdf"},
        )

        fields = {fact["field"]: fact for fact in workflow["facts"]}
        self.assertEqual(fields["trading_name"]["value"], "Jenny's Early Learning Norlane")
        self.assertEqual(fields["address"]["source"]["page"], 2)
        self.assertEqual(fields["suburb"]["value"], "Norlane")
        self.assertEqual(fields["postcode"]["value"], "3214")
        self.assertEqual(fields["postcode"]["provenance"], "found")
        self.assertEqual(fields["postcode"]["source_type"], "supplemental_doc")
        self.assertNotIn("postcode", [str(field).lower() for field in workflow["missing_fields"]])

    def test_manual_evidence_notes_are_low_confidence_and_redacted(self):
        text = main.manual_evidence_text([
            main.ManualEvidenceNote(
                source_label="Payroll diligence item",
                diligence_item_id="item-1",
                status="received",
                category="financials",
                question="Confirm payroll",
                notes="Broker says payroll is 620000. API key sk-ant-secret should not be present elsewhere.",
            )
        ])

        self.assertIn("SOURCE_TYPE: manual_user_note", text)
        self.assertIn("CONFIDENCE: low", text)
        self.assertIn("STATUS: received", text)
        self.assertNotIn("sk-ant-secret", text)
        self.assertIn("[redacted_api_key]", text)
        self.assertIn("Do not silently override source-backed document values", text)

    def test_manual_context_fact_is_low_trust_review_context(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["note"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 1000000, "ebitda": 200000, "total_labour_cost": 550000}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== Manual diligence notes (manual_user_note) ===\n"
                "--- Manual Evidence 1 ---\n"
                "SOURCE_TYPE: manual_user_note\n"
                "STATUS: received\n"
                "QUESTION: Director retention\n"
                "NOTES: Vendor says the director will stay after settlement.\n"
            ),
            source_files=["Manual diligence notes"],
            file_classes={"Manual diligence notes": "manual_user_note"},
        )

        manual = [fact for fact in workflow["facts"] if fact["provenance"] == "manual_context"]
        self.assertTrue(manual)
        self.assertEqual(manual[0]["trust"], "low")
        self.assertEqual(manual[0]["underwriting_use"], "review_required")
        self.assertTrue(workflow["evidence_readiness"]["manual_context"])
        self.assertTrue(workflow["partner_judgement_prompts"])

    def test_ytd_financial_period_is_partial_review_required(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["management.xlsx"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== management.xlsx (pl_excel) ===\n"
                "SHEET: Management Accounts YTD May25-Feb26\n"
                "D23=Total revenue YTD | E23=900000\n"
                "D30=Total Employment Costs YTD | E30=520000\n"
                "D41=EBITDA YTD | E41=140000\n"
            ),
            source_files=["management.xlsx"],
            file_classes={"management.xlsx": "pl_excel"},
        )

        revenue = next(f for f in workflow["facts"] if f["field"] == "revenue")
        self.assertEqual(revenue["period"]["coverage_status"], "partial")
        self.assertEqual(revenue["underwriting_use"], "review_required")
        self.assertIn("YTD", revenue["period"]["coverage_reason"])

    def test_template_payroll_is_excluded_from_underwriting(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["staffing-template.xlsx"], "missing_fields": ["payroll"]},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 1000000, "ebitda": 200000}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== staffing-template.xlsx (payroll_excel) ===\n"
                "SHEET: Staffing Template\n"
                "A1=Template payroll assumption | B1=Payroll | C1=620000\n"
            ),
            source_files=["staffing-template.xlsx"],
            file_classes={"staffing-template.xlsx": "payroll_excel"},
        )

        payroll = next(f for f in workflow["facts"] if f["field"] == "payroll_labour_cost")
        self.assertEqual(payroll["source_quality"], "template_or_forecast")
        self.assertEqual(payroll["underwriting_use"], "excluded")
        self.assertEqual(workflow["valuation_gate"]["status"], "blocked")

    def test_monthly_occupancy_does_not_unlock_weekly_averages(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["occupancy.xlsx"], "missing_fields": ["occupancy_history"]},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 1000000, "ebitda": 200000, "total_labour_cost": 550000}},
                "occupancy": {},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== occupancy.xlsx (occupancy_excel) ===\n"
                "SHEET: Output Occupancy\n"
                "A1=Month | B1=Occupancy\n"
                "A2=Jan | B2=0.80\n"
                "A3=Feb | B3=0.82\n"
                "A4=Mar | B4=0.84\n"
                "A5=Apr | B5=0.86\n"
            ),
            source_files=["occupancy.xlsx"],
            file_classes={"occupancy.xlsx": "occupancy_excel"},
        )

        fields = {f["field"]: f for f in workflow["facts"]}
        self.assertIn("monthly_avg_occupancy_pct", fields)
        self.assertIsNone(fields["avg_13wk_occupancy_pct"]["value"])
        self.assertEqual(fields["avg_13wk_occupancy_pct"]["underwriting_use"], "blocked")
        self.assertIn("13 weekly observations", fields["avg_13wk_occupancy_pct"]["period"]["coverage_reason"])

    def test_canonical_facts_prevent_legacy_financial_contradictions(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["im.pdf", "databook.xlsx"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 163000, "ebitda": 336000, "total_labour_cost": 30000}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== databook.xlsx (pl_excel) ===\n"
                "SHEET: Adjusted Actuals\n"
                "D23=Total revenue FY25 | E23=2374250\n"
                "D30=Total Employment Costs FY25 | E30=1289062\n"
                "D41=EBITDA FY25 | E41=347218\n"
            ),
            source_files=["im.pdf", "databook.xlsx"],
            file_classes={"im.pdf": "im_pdf", "databook.xlsx": "pl_excel"},
        )

        canonical = workflow["canonical_facts"]
        self.assertEqual(canonical["revenue"]["value"], 2374250)
        self.assertEqual(canonical["payroll_labour_cost"]["value"], 1289062)
        self.assertNotEqual(canonical["revenue"]["value"], 163000)
        self.assertEqual(canonical["revenue"]["status"], "conflicting")

    def test_valuation_gate_summary_uses_underwriting_use_not_present(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["management.xlsx"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== management.xlsx (pl_excel) ===\n"
                "SHEET: Management Accounts YTD May25-Feb26\n"
                "D23=Total revenue YTD | E23=900000\n"
                "D30=Total Employment Costs YTD | E30=520000\n"
                "D41=EBITDA YTD | E41=140000\n"
            ),
            source_files=["management.xlsx"],
            file_classes={"management.xlsx": "pl_excel"},
        )

        rows = {row["field"]: row for row in workflow["valuation_gate_summary"]["rows"]}
        self.assertEqual(rows["revenue"]["evidence"], "found")
        self.assertEqual(rows["revenue"]["underwriting_use"], "review_required")
        self.assertIn("review", rows["revenue"]["underwriting_use"])

    def test_market_audit_sanitizes_provider_errors_and_normalizes_fallback(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["im.pdf"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 1000000, "ebitda": 200000, "total_labour_cost": 550000}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
                "_market_audit": {
                    "warnings": ["{'code': '42703', 'message': 'column acecqa_centres.service_approval_number does not exist'}"],
                    "competitor_supply": {
                        "source": "unavailable",
                        "confidence": "low",
                        "competitor_count": 0,
                        "total_licensed_places": 74,
                        "warnings": ["postgres column acecqa_centres.service_approval_number does not exist"],
                        "compared_to_postcode": {"competitor_count": 6, "total_licensed_places": 736, "edr": 0.92},
                    },
                },
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text="",
            source_files=["im.pdf"],
            file_classes={"im.pdf": "im_pdf"},
        )

        audit = workflow["market_audit"]
        warnings = " ".join(audit["warnings"] + audit["competitor_supply"]["warnings"])
        self.assertNotIn("42703", warnings)
        self.assertNotIn("column acecqa", warnings.lower())
        self.assertIn("Competitor lookup failed", warnings)
        self.assertIsNone(audit["competitor_supply"]["competitor_count"])
        self.assertIsNone(audit["competitor_supply"]["total_licensed_places"])
        self.assertEqual(audit["competitor_supply"]["compared_to_postcode"]["competitor_count"], 6)

    def test_missing_occupancy_fields_collapse_to_document_request(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["im.pdf"], "missing_fields": ["latest_week_occupancy_pct", "avg_4wk_occupancy_pct", "avg_13wk_occupancy_pct", "lease_expiry"]},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {"revenue": 1000000, "ebitda": 200000, "total_labour_cost": 550000}},
                "occupancy": {},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text="",
            source_files=["im.pdf"],
            file_classes={"im.pdf": "im_pdf"},
        )

        requests = [item["question"] for item in workflow["diligence_checklist"]]
        joined = " ".join(requests)
        self.assertIn("Upload occupancy/utilisation export", joined)
        self.assertIn("Upload executed lease", joined)
        self.assertNotIn("avg_4wk_occupancy_pct", joined)

    def test_evidence_quality_not_high_when_underwriting_requires_review(self):
        workflow = build_structured_deal_intelligence(
            extracted={
                "meta": {"source_files": ["management.xlsx"], "missing_fields": []},
                "centre": {"name": "Synthetic Childcare"},
                "financials": {"fy25": {}},
                "occupancy": {"avg_4wk_pct": 80, "avg_13wk_pct": 81},
                "lease": {},
                "fees": {},
            },
            scored={"centre_name": "Synthetic Childcare"},
            combined_text=(
                "=== management.xlsx (pl_excel) ===\n"
                "SHEET: Management Accounts YTD May25-Feb26\n"
                "D23=Total revenue YTD | E23=900000\n"
                "D30=Total Employment Costs YTD | E30=520000\n"
                "D41=EBITDA YTD | E41=140000\n"
            ),
            source_files=["management.xlsx"],
            file_classes={"management.xlsx": "pl_excel"},
        )

        self.assertNotEqual(workflow["evidence_quality"]["evidence_quality"], "High")
        self.assertEqual(workflow["evidence_quality"]["underwriting_reliability"], "Review required")


if __name__ == "__main__":
    unittest.main()
